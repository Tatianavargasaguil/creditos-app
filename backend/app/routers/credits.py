from datetime import date, datetime, time
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import Response, StreamingResponse
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, joinedload

from app import models, schemas
from app.database import get_db
from app.email_service import send_alert_email
from app.security import get_current_user, require_admin, require_credit_operator

router = APIRouter(prefix="/credits", tags=["credits"])


def _next_reference(db: Session) -> str:
    next_number = (db.query(func.coalesce(func.max(models.CreditRequest.id), 0)).scalar() or 0) + 1
    while True:
        reference = f"CRE-{next_number:05d}"
        exists = db.query(models.CreditRequest.id).filter(models.CreditRequest.reference == reference).first()
        if not exists:
            return reference
        next_number += 1


def _default_stage(db: Session) -> models.Stage:
    stage = db.query(models.Stage).filter(models.Stage.code == "viabilidad").first()
    if not stage:
        raise HTTPException(status_code=500, detail="No existe la etapa Viabilidad")
    return stage


def _credit_query(db: Session):
    return (
        db.query(models.CreditRequest)
        .options(
            joinedload(models.CreditRequest.stage),
            joinedload(models.CreditRequest.viability_bank),
            joinedload(models.CreditRequest.selected_bank),
            joinedload(models.CreditRequest.disbursement_bank),
            joinedload(models.CreditRequest.bank_lines).joinedload(models.CreditBankLine.bank),
            joinedload(models.CreditRequest.documents),
            joinedload(models.CreditRequest.alerts),
            joinedload(models.CreditRequest.history),
        )
    )

def _advisor_credit_response(credit: models.CreditRequest) -> schemas.CreditRequestRead:
    data = schemas.CreditRequestRead.model_validate(credit)
    return data.model_copy(update={
        "phone": None,
        "brand": None,
        "line": None,
        "model": None,
        "sale_price": 0,
        "down_payment": 0,
        "proforma_invoice_ref": None,
        "final_invoice_ref": None,
        "viability_bank_id": None,
        "selected_bank_id": None,
        "disbursement_bank_id": None,
        "approval_conditions": None,
        "observations": None,
        "rejection_reason": None,
        "ok_runt": False,
        "ok_runt_at": None,
        "runt_observation": None,
        "insured_ok": False,
        "policy_issued": False,
        "insurance_company": None,
        "policy_observation": None,
        "disbursed_value": 0,
        "ownership_card_issued": False,
        "ownership_card_delivery_date": None,
        "viability_bank": None,
        "selected_bank": None,
        "disbursement_bank": None,
        "documents": [],
        "alerts": [],
        "history": [],
    })

HISTORY_FIELD_LABELS = {
    "customer_name": "Cliente",
    "document_type": "Tipo de documento",
    "document_number": "Numero de documento",
    "phone": "Celular",
    "plate": "Placa",
    "vin": "VIN",
    "brand": "Marca",
    "line": "Linea",
    "model": "Modelo",
    "advisor_name": "Asesor",
    "showroom": "Vitrina",
    "business_type": "Tipo de negocio",
    "sale_price": "Precio de venta",
    "down_payment": "Cuota inicial",
    "stage_id": "Etapa",
    "viability_bank_id": "Banco de viabilidad",
    "selected_bank_id": "Banco firmado",
    "disbursement_bank_id": "Banco de desembolso",
    "rejection_reason": "Motivo de rechazo",
    "ok_runt": "OK RUNT / Preinscripcion de prenda",
    "runt_observation": "Observacion RUNT",
    "policy_issued": "Poliza emitida",
    "insurance_company": "Entidad aseguradora",
    "policy_observation": "Observacion de poliza",
    "disbursed_value": "Valor desembolsado",
    "ownership_card_issued": "Tarjeta de propiedad emitida",
    "ownership_card_delivery_date": "Fecha entrega tarjeta de propiedad",
    "proforma_invoice_ref": "Factura proforma",
    "final_invoice_ref": "Factura definitiva",
    "approval_conditions": "Condiciones de aprobacion",
    "observations": "Observaciones",
}


def _humanize_history_detail(detail: str | None) -> str:
    if not detail:
        return "-"

    clean_detail = detail.strip()
    if clean_detail.startswith("Etapa anterior ID "):
        return clean_detail.replace("Etapa anterior ID", "Etapa anterior").replace(", nueva ID", ", nueva etapa")

    parts = [part.strip() for part in clean_detail.split(",") if part.strip()]
    should_translate = any(part in HISTORY_FIELD_LABELS or "_" in part for part in parts)
    if not should_translate:
        return clean_detail

    return ", ".join(HISTORY_FIELD_LABELS.get(part, part.replace("_", " ").title()) for part in parts)

def _add_history(db: Session, credit_id: int, action: str, detail: str | None = None, actor: str = "Sistema") -> None:
    db.add(models.CreditHistory(credit_id=credit_id, actor=actor, action=action, detail=detail))


@router.get("", response_model=list[schemas.CreditRequestRead])
def list_credits(
    search: str | None = Query(default=None),
    stage_code: str | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    query = _credit_query(db)
    clean_search = (search or "").strip()
    if current_user.role == models.UserRole.advisor:
        if not clean_search:
            return []
        like = f"%{clean_search}%"
        query = query.filter(
            or_(
                models.CreditRequest.document_number.ilike(like),
                models.CreditRequest.plate.ilike(like),
                models.CreditRequest.vin.ilike(like),
            )
        )
    elif clean_search:
        like = f"%{clean_search}%"
        query = query.filter(
            or_(
                models.CreditRequest.reference.ilike(like),
                models.CreditRequest.customer_name.ilike(like),
                models.CreditRequest.document_number.ilike(like),
                models.CreditRequest.plate.ilike(like),
                models.CreditRequest.vin.ilike(like),
            )
        )
    if stage_code:
        query = query.join(models.Stage).filter(models.Stage.code == stage_code)
    credits = query.order_by(models.CreditRequest.created_at.desc()).all()
    if current_user.role == models.UserRole.advisor:
        return [_advisor_credit_response(credit) for credit in credits]
    return credits


@router.post("", response_model=schemas.CreditRequestRead, status_code=status.HTTP_201_CREATED)
def create_credit(
    payload: schemas.CreditRequestCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_credit_operator),
):
    stage_id = payload.stage_id or _default_stage(db).id
    values = payload.model_dump(exclude={"stage_id"})
    credit = models.CreditRequest(reference=_next_reference(db), stage_id=stage_id, **values)
    db.add(credit)
    db.flush()
    _add_history(db, credit.id, "Credito creado", "Registro inicial de solicitud", current_user.full_name)
    db.commit()
    return _credit_query(db).filter(models.CreditRequest.id == credit.id).one()


@router.get("/reports/summary", response_model=schemas.DashboardSummary)
def summary(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if current_user.role == models.UserRole.advisor:
        return schemas.DashboardSummary(total_requests=0, active_requests=0, approved_requests=0, disbursed_value=0, by_stage={}, by_selected_bank={})
    total = db.query(models.CreditRequest).count()
    active = (
        db.query(models.CreditRequest)
        .join(models.Stage)
        .filter(models.Stage.code.notin_(["legalizacion", "desasignado"]))
        .count()
    )
    approved = db.query(models.CreditRequest).join(models.Stage).filter(models.Stage.code == "aprobado").count()
    disbursed_value = db.query(func.coalesce(func.sum(models.CreditRequest.disbursed_value), 0)).scalar() or 0

    by_stage_rows = (
        db.query(models.Stage.name, func.count(models.CreditRequest.id))
        .join(models.CreditRequest, models.CreditRequest.stage_id == models.Stage.id, isouter=True)
        .group_by(models.Stage.name, models.Stage.sequence)
        .order_by(models.Stage.sequence)
        .all()
    )
    by_bank_rows = (
        db.query(models.Bank.name, func.count(models.CreditRequest.id))
        .join(models.CreditRequest, models.CreditRequest.selected_bank_id == models.Bank.id)
        .group_by(models.Bank.name)
        .all()
    )
    return schemas.DashboardSummary(
        total_requests=total,
        active_requests=active,
        approved_requests=approved,
        disbursed_value=float(disbursed_value),
        by_stage={name: count for name, count in by_stage_rows},
        by_selected_bank={name: count for name, count in by_bank_rows},
    )


@router.get("/reports/detail", response_model=list[schemas.CreditRequestRead])
def detail_report(
    search: str | None = Query(default=None),
    stage_id: int | None = Query(default=None),
    bank_id: int | None = Query(default=None),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    db: Session = Depends(get_db),
    _current_user: models.User = Depends(require_credit_operator),
):
    query = _credit_query(db)
    if search:
        like = f"%{search}%"
        query = query.filter(
            or_(
                models.CreditRequest.reference.ilike(like),
                models.CreditRequest.customer_name.ilike(like),
                models.CreditRequest.document_number.ilike(like),
                models.CreditRequest.plate.ilike(like),
                models.CreditRequest.vin.ilike(like),
                models.CreditRequest.advisor_name.ilike(like),
                models.CreditRequest.showroom.ilike(like),
            )
        )
    if stage_id:
        query = query.filter(models.CreditRequest.stage_id == stage_id)
    if bank_id:
        query = query.filter(
            or_(
                models.CreditRequest.viability_bank_id == bank_id,
                models.CreditRequest.selected_bank_id == bank_id,
                models.CreditRequest.disbursement_bank_id == bank_id,
                models.CreditRequest.bank_lines.any(models.CreditBankLine.bank_id == bank_id),
            )
        )
    if date_from:
        query = query.filter(models.CreditRequest.created_at >= datetime.combine(date_from, time.min))
    if date_to:
        query = query.filter(models.CreditRequest.created_at <= datetime.combine(date_to, time.max))
    return query.order_by(models.CreditRequest.created_at.desc()).all()


def _apply_report_filters(query, search, stage_id, bank_id, date_from, date_to):
    if search:
        like = f"%{search}%"
        query = query.filter(
            or_(
                models.CreditRequest.reference.ilike(like),
                models.CreditRequest.customer_name.ilike(like),
                models.CreditRequest.document_number.ilike(like),
                models.CreditRequest.plate.ilike(like),
                models.CreditRequest.vin.ilike(like),
                models.CreditRequest.advisor_name.ilike(like),
                models.CreditRequest.showroom.ilike(like),
            )
        )
    if stage_id:
        query = query.filter(models.CreditRequest.stage_id == stage_id)
    if bank_id:
        query = query.filter(
            or_(
                models.CreditRequest.viability_bank_id == bank_id,
                models.CreditRequest.selected_bank_id == bank_id,
                models.CreditRequest.disbursement_bank_id == bank_id,
                models.CreditRequest.bank_lines.any(models.CreditBankLine.bank_id == bank_id),
            )
        )
    if date_from:
        query = query.filter(models.CreditRequest.created_at >= datetime.combine(date_from, time.min))
    if date_to:
        query = query.filter(models.CreditRequest.created_at <= datetime.combine(date_to, time.max))
    return query


def _join_lines(values: list[str]) -> str:
    return "\n".join(value for value in values if value)


BANK_TYPE_LABELS = {
    "viabilidad": "Viabilidad",
    "estudio": "Estudio",
    "aprobacion": "Aprobacion",
    "rechazo": "Rechazo",
    "desembolso": "Desembolso",
}


BANK_STATUS_LABELS = {
    "pendiente": "Pendiente",
    "radicado": "Radicado",
    "aprobado": "Aprobado",
    "negado": "Negado",
    "desembolsado": "Desembolsado",
}


def _bank_names_by_type(credit: models.CreditRequest, bank_type: str) -> str:
    names: list[str] = []
    for line in credit.bank_lines:
        if line.type != bank_type or not line.bank:
            continue
        if line.bank.name not in names:
            names.append(line.bank.name)
    return ", ".join(names)


def _bank_details_by_type(credit: models.CreditRequest, bank_type: str) -> str:
    values: list[str] = []
    for line in credit.bank_lines:
        if line.type != bank_type:
            continue
        bank_name = line.bank.name if line.bank else f"Banco ID {line.bank_id}"
        notes = line.conditions or line.rejection_reason or "-"
        values.append(
            " | ".join(
                [
                    bank_name,
                    BANK_STATUS_LABELS.get(line.status, line.status),
                    f"radicado: {line.filed_at or '-'}",
                    f"respuesta: {line.answered_at or '-'}",
                    f"observacion: {notes}",
                ]
            )
        )
    return _join_lines(values)


@router.get("/reports/excel")
def excel_report(
    search: str | None = Query(default=None),
    stage_id: int | None = Query(default=None),
    bank_id: int | None = Query(default=None),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    db: Session = Depends(get_db),
    _current_user: models.User = Depends(require_credit_operator),
):
    query = _apply_report_filters(_credit_query(db), search, stage_id, bank_id, date_from, date_to)
    credits = query.order_by(models.CreditRequest.created_at.desc()).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Solicitudes"
    headers = [
        "Referencia",
        "Cliente",
        "Tipo documento",
        "Numero documento",
        "Celular",
        "Placa",
        "VIN",
        "Vehiculo",
        "Asesor",
        "Vitrina",
        "Tipo negocio",
        "Etapa actual",
        "Fecha creacion",
        "Inicio etapa actual",
        "Valor venta",
        "Cuota inicial",
        "Valor financiado",
        "Valor desembolsado",
        "Bancos viabilidad",
        "Detalle viabilidad",
        "Bancos estudio",
        "Detalle estudio",
        "Bancos aprobados",
        "Detalle aprobacion",
        "Banco firmado",
        "Banco desembolso",
        "Detalle desembolso",
        "Motivo rechazo",
        "OK RUNT",
        "Observacion RUNT",
        "Asegurado OK",
        "Poliza emitida",
        "Entidad aseguradora",
        "Observacion poliza",
        "Tarjeta propiedad emitida",
        "Fecha entrega tarjeta",
        "Documentos",
        "Alertas",
        "Historial / etapas",
        "Observaciones",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1967D2")

    for credit in credits:
        documents = _join_lines([
            f"{document.name} | {document.type} | {document.file_name or '-'} | {document.file_size or 0} bytes"
            for document in credit.documents
        ])
        alerts = _join_lines([
            f"{alert.type} | {alert.status} | {alert.message} | correo: {alert.email_to or '-'} | enviado: {'si' if alert.email_sent else 'no'}"
            for alert in credit.alerts
        ])
        history = _join_lines([
            f"{item.created_at} | {item.actor} | {item.action} | {_humanize_history_detail(item.detail)}"
            for item in credit.history
        ])
        sheet.append([
            credit.reference,
            credit.customer_name,
            credit.document_type,
            credit.document_number,
            credit.phone,
            credit.plate,
            credit.vin,
            f"{credit.brand or ''} {credit.line or ''} {credit.model or ''}".strip(),
            credit.advisor_name,
            credit.showroom,
            credit.business_type,
            credit.stage.name if credit.stage else "",
            credit.created_at,
            credit.stage_started_at,
            float(credit.sale_price or 0),
            float(credit.down_payment or 0),
            credit.financed_value,
            float(credit.disbursed_value or 0),
            _bank_names_by_type(credit, "viabilidad") or (credit.viability_bank.name if credit.viability_bank else ""),
            _bank_details_by_type(credit, "viabilidad"),
            _bank_names_by_type(credit, "estudio"),
            _bank_details_by_type(credit, "estudio"),
            _bank_names_by_type(credit, "aprobacion"),
            _bank_details_by_type(credit, "aprobacion"),
            credit.selected_bank.name if credit.selected_bank else "",
            _bank_names_by_type(credit, "desembolso") or (credit.disbursement_bank.name if credit.disbursement_bank else ""),
            _bank_details_by_type(credit, "desembolso"),
            credit.rejection_reason,
            "Si" if credit.ok_runt else "No",
            credit.runt_observation,
            "Si" if credit.insured_ok else "No",
            "Si" if credit.policy_issued else "No",
            credit.insurance_company,
            credit.policy_observation,
            "Si" if credit.ownership_card_issued else "No",
            credit.ownership_card_delivery_date,
            documents,
            alerts,
            history,
            credit.observations,
        ])

    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 60)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"reporte_creditos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{credit_id}", response_model=schemas.CreditRequestRead)
def get_credit(credit_id: int, db: Session = Depends(get_db), _current_user: models.User = Depends(require_credit_operator)):
    credit = _credit_query(db).filter(models.CreditRequest.id == credit_id).first()
    if not credit:
        raise HTTPException(status_code=404, detail="Credito no encontrado")
    return credit


@router.patch("/{credit_id}", response_model=schemas.CreditRequestRead)
def update_credit(
    credit_id: int,
    payload: schemas.CreditRequestUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_credit_operator),
):
    credit = db.get(models.CreditRequest, credit_id)
    if not credit:
        raise HTTPException(status_code=404, detail="Credito no encontrado")

    values = payload.model_dump(exclude_unset=True)
    old_stage_id = credit.stage_id
    for key, value in values.items():
        setattr(credit, key, value)
    if "stage_id" in values and values["stage_id"] != old_stage_id:
        credit.stage_started_at = datetime.utcnow()
        _add_history(
            db,
            credit.id,
            "Cambio de etapa",
            f"Etapa anterior ID {old_stage_id}, nueva ID {values['stage_id']}",
            current_user.full_name,
        )
    else:
        _add_history(db, credit.id, "Credito actualizado", _humanize_history_detail(", ".join(values.keys())), current_user.full_name)
    db.commit()
    return _credit_query(db).filter(models.CreditRequest.id == credit.id).one()


@router.delete("/{credit_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_credit(
    credit_id: int,
    db: Session = Depends(get_db),
    _current_user: models.User = Depends(require_admin),
):
    credit = db.get(models.CreditRequest, credit_id)
    if not credit:
        raise HTTPException(status_code=404, detail="Credito no encontrado")

    db.query(models.Notification).filter(models.Notification.credit_id == credit_id).update(
        {models.Notification.credit_id: None},
        synchronize_session=False,
    )
    db.delete(credit)
    db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.post("/{credit_id}/bank-lines", response_model=schemas.BankLineRead, status_code=status.HTTP_201_CREATED)
def add_bank_line(
    credit_id: int,
    payload: schemas.BankLineCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_credit_operator),
):
    if not db.get(models.CreditRequest, credit_id):
        raise HTTPException(status_code=404, detail="Credito no encontrado")
    line = models.CreditBankLine(credit_id=credit_id, **payload.model_dump())
    db.add(line)
    db.flush()
    _add_history(db, credit_id, "Banco agregado", f"Banco ID {payload.bank_id} en estado {payload.status}", current_user.full_name)
    db.commit()
    db.refresh(line)
    return line


@router.delete("/{credit_id}/bank-lines/{line_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_bank_line(
    credit_id: int,
    line_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_credit_operator),
):
    line = (
        db.query(models.CreditBankLine)
        .options(joinedload(models.CreditBankLine.bank))
        .filter(models.CreditBankLine.credit_id == credit_id, models.CreditBankLine.id == line_id)
        .first()
    )
    if not line:
        raise HTTPException(status_code=404, detail="Banco no encontrado en el credito")

    bank_name = line.bank.name if line.bank else f"Banco ID {line.bank_id}"
    detail = f"{bank_name} | {BANK_TYPE_LABELS.get(line.type, line.type)} | {BANK_STATUS_LABELS.get(line.status, line.status)}"
    db.delete(line)
    _add_history(db, credit_id, "Banco eliminado", detail, current_user.full_name)
    db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.post("/{credit_id}/documents", response_model=schemas.DocumentRead, status_code=status.HTTP_201_CREATED)
def add_document(
    credit_id: int,
    payload: schemas.DocumentCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_credit_operator),
):
    if not db.get(models.CreditRequest, credit_id):
        raise HTTPException(status_code=404, detail="Credito no encontrado")
    document = models.CreditDocument(credit_id=credit_id, **payload.model_dump())
    db.add(document)
    db.flush()
    _add_history(db, credit_id, "Documento agregado", payload.name, current_user.full_name)
    db.commit()
    db.refresh(document)
    return document


@router.post("/{credit_id}/documents/upload", response_model=schemas.DocumentRead, status_code=status.HTTP_201_CREATED)
def upload_document(
    credit_id: int,
    name: str = Form(...),
    type: str = Form("otro"),
    observation: str | None = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_credit_operator),
):
    if not db.get(models.CreditRequest, credit_id):
        raise HTTPException(status_code=404, detail="Credito no encontrado")

    file_data = file.file.read()
    safe_name = file.filename or "documento"

    document = models.CreditDocument(
        credit_id=credit_id,
        name=name,
        type=type,
        file_name=safe_name,
        file_url=f"/api/credits/{credit_id}/documents/download",
        mime_type=file.content_type or "application/octet-stream",
        file_size=len(file_data),
        file_data=file_data,
        observation=observation,
    )
    db.add(document)
    db.flush()
    _add_history(db, credit_id, "Documento cargado", safe_name, current_user.full_name)
    db.commit()
    db.refresh(document)
    return document


@router.get("/{credit_id}/documents/{document_id}/download")
def download_document(
    credit_id: int,
    document_id: int,
    db: Session = Depends(get_db),
    _current_user: models.User = Depends(require_credit_operator),
):
    document = (
        db.query(models.CreditDocument)
        .filter(models.CreditDocument.credit_id == credit_id, models.CreditDocument.id == document_id)
        .first()
    )
    if not document or not document.file_data:
        raise HTTPException(status_code=404, detail="Documento no encontrado")

    filename = document.file_name or document.name
    return Response(
        content=document.file_data,
        media_type=document.mime_type or "application/octet-stream",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


@router.delete("/{credit_id}/documents/{document_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_document(
    credit_id: int,
    document_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    document = (
        db.query(models.CreditDocument)
        .filter(models.CreditDocument.credit_id == credit_id, models.CreditDocument.id == document_id)
        .first()
    )
    if not document:
        raise HTTPException(status_code=404, detail="Documento no encontrado")

    detail = document.file_name or document.name
    db.delete(document)
    _add_history(db, credit_id, "Documento eliminado", detail, current_user.full_name)
    db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.post("/{credit_id}/alerts", response_model=schemas.AlertRead, status_code=status.HTTP_201_CREATED)
def add_alert(
    credit_id: int,
    type: str = Form("otro"),
    message: str = Form(...),
    recipients: str = Form("equipo_creditos"),
    email_to: str | None = Form(None),
    selected_document_ids: list[int] = Form([]),
    file: UploadFile | None = File(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_credit_operator),
):
    if not db.get(models.CreditRequest, credit_id):
        raise HTTPException(status_code=404, detail="Credito no encontrado")
    alert = models.CreditAlert(
        credit_id=credit_id,
        type=type,
        message=message,
        recipients=recipients,
        email_to=email_to,
        scheduled_at=datetime.utcnow(),
    )
    db.add(alert)
    db.flush()
    attachments: list[tuple[str, str, bytes]] = []
    attached_names: list[str] = []

    if selected_document_ids:
        documents = (
            db.query(models.CreditDocument)
            .filter(
                models.CreditDocument.credit_id == credit_id,
                models.CreditDocument.id.in_(selected_document_ids),
                models.CreditDocument.file_data.isnot(None),
            )
            .all()
        )
        for document in documents:
            filename = document.file_name or document.name
            attachments.append((filename, document.mime_type or "application/octet-stream", document.file_data or b""))
            attached_names.append(filename)

    if file:
        file_data = file.file.read()
        safe_name = file.filename or "adjunto_alerta"
        attachments.append((safe_name, file.content_type or "application/octet-stream", file_data))
        attached_names.append(safe_name)
        document = models.CreditDocument(
            credit_id=credit_id,
            name=safe_name,
            type="alerta",
            file_name=safe_name,
            file_url=f"/api/credits/{credit_id}/documents/download",
            mime_type=file.content_type or "application/octet-stream",
            file_size=len(file_data),
            file_data=file_data,
            observation=f"Adjunto enviado en alerta {type}",
        )
        db.add(document)
        db.flush()
        _add_history(db, credit_id, "Documento cargado", safe_name, current_user.full_name)

    if alert.email_to:
        credit = _credit_query(db).filter(models.CreditRequest.id == credit_id).one()
        try:
            send_alert_email(credit, alert, attachments)
            alert.status = models.AlertStatus.enviado
            alert.email_sent = True
            alert.email_sent_at = datetime.utcnow()
            alert.sent_at = alert.email_sent_at
            alert.email_error = None
            alert.error_message = None
        except Exception as error:
            alert.status = models.AlertStatus.fallido
            alert.email_sent = False
            alert.email_error = str(error)
            alert.error_message = str(error)
    detail = message
    if attached_names:
        detail = f"{message} | Adjuntos: {', '.join(attached_names)}"
    _add_history(db, credit_id, "Alerta creada", detail, current_user.full_name)
    db.commit()
    db.refresh(alert)
    return alert





